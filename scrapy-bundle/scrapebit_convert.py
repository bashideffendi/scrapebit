"""
Convert scrapy output_*.json ke Excel / CSV.

Excel: 2 sheets per file
  - "Annual" — metrics dengan period taun (e.g., 2020, 2025)
  - "Quarterly" — metrics dengan period kuartal (e.g., 2025-Q1, 2026-Q1)

Tiap sheet, metric di-group + sort by section: IS → BS → BS-Bank → CF → Ratios.

CSV: long format flat (ticker,name,sector,section,metric,period,value).

Usage:
    python scrapebit_convert.py --input output.json --output output.xlsx --format excel
    python scrapebit_convert.py --input output.json --output output.csv  --format csv
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path


META_FIELDS = {"tick", "name", "sector", "subsector"}

# Field-field non-time-series — skip dari export wide-format
NON_TIMESERIES_FIELDS = {
    "PricePerformance",
    "DividendEvents",
    "EmittenStats",
    "EmittenInfo",
    "KeystatsRaw",
    "ClosureItems",
    "DailyChart",
}

# Classifier metric → section. Used for grouping + sort order di Excel.
# Order matters: 1 = Income Statement, 2 = Balance Sheet, dst.
METRIC_SECTIONS = {
    # ── 1. Income Statement ─────────────────────────────────
    "Revenue": (1, "Income Statement"),
    "Interest Income": (1, "Income Statement"),
    "Fee & Commission Income": (1, "Income Statement"),
    "Beban Pkk Penjualan": (1, "Income Statement"),
    "Interest Expense": (1, "Income Statement"),
    "Gross Profit": (1, "Income Statement"),
    "Biaya Operasional": (1, "Income Statement"),
    "Employee Expense": (1, "Income Statement"),
    "G&A Expense": (1, "Income Statement"),
    "Loan Loss Provision": (1, "Income Statement"),
    "Selling Expense": (1, "Income Statement"),
    "EBIT": (1, "Income Statement"),
    "Other Income/Expense": (1, "Income Statement"),
    "Biaya Keuangan": (1, "Income Statement"),
    "Pre-Tax Income": (1, "Income Statement"),
    "Current Tax": (1, "Income Statement"),
    "Deferred Tax": (1, "Income Statement"),
    "Income Tax Expense": (1, "Income Statement"),
    "Net Income Total": (1, "Income Statement"),
    "Net Profit": (1, "Income Statement"),
    "Minority Interest Income": (1, "Income Statement"),
    "OCI": (1, "Income Statement"),
    "Comprehensive Income": (1, "Income Statement"),
    # ── 2. Balance Sheet (Assets) ───────────────────────────
    "Cash": (2, "Balance Sheet"),
    "Piutang": (2, "Balance Sheet"),
    "Persediaan": (2, "Balance Sheet"),
    "Aset Lancar": (2, "Balance Sheet"),
    "Aset Tidak Lancar": (2, "Balance Sheet"),
    "Total Assets": (2, "Balance Sheet"),
    "Fixed Assets": (2, "Balance Sheet"),
    "Intangible Assets": (2, "Balance Sheet"),
    "Deferred Tax Assets": (2, "Balance Sheet"),
    "Consumer Financing Receivables": (2, "Balance Sheet"),
    "Government Bonds": (2, "Balance Sheet"),
    # ── 3. Balance Sheet (Bank-specific) ────────────────────
    "Customer Deposits": (3, "BS - Bank"),
    "BI Current Account": (3, "BS - Bank"),
    "BI Placements": (3, "BS - Bank"),
    "Loans Given": (3, "BS - Bank"),
    "Interbank Deposits": (3, "BS - Bank"),
    "Other Bank Liabilities": (3, "BS - Bank"),
    "Acceptance Liabilities": (3, "BS - Bank"),
    "Issued Securities": (3, "BS - Bank"),
    "Borrowings": (3, "BS - Bank"),
    "Subordinated Bonds": (3, "BS - Bank"),
    "Syirkah Temporer Funds": (3, "BS - Bank"),
    "Immediate Liabilities": (3, "BS - Bank"),
    # ── 4. Balance Sheet (Liab + Equity) ────────────────────
    "Liab J Pndk": (4, "BS - Liab/Equity"),
    "Liab J Pnjg": (4, "BS - Liab/Equity"),
    "Total Liabilitas": (4, "BS - Liab/Equity"),
    "Deferred Tax Liabilities": (4, "BS - Liab/Equity"),
    "Ekuitas": (4, "BS - Liab/Equity"),
    "Total Equity": (4, "BS - Liab/Equity"),
    "Total Liab + Equity": (4, "BS - Liab/Equity"),
    "Share Capital": (4, "BS - Liab/Equity"),
    "Additional Paid-in Capital": (4, "BS - Liab/Equity"),
    "Retained Earnings": (4, "BS - Liab/Equity"),
    "Minority Interest": (4, "BS - Liab/Equity"),
    "Shares Outstanding (BS)": (4, "BS - Liab/Equity"),
    # ── 5. Cash Flow ────────────────────────────────────────
    "CFO": (5, "Cash Flow"),
    "Cash flow from investor": (5, "Cash Flow"),
    "Cash flow from finance": (5, "Cash Flow"),
    "Net Change in Cash": (5, "Cash Flow"),
    "Cash Begin Period": (5, "Cash Flow"),
    "Cash End Period": (5, "Cash Flow"),
    "Cash Interest Received": (5, "Cash Flow"),
    "Cash Interest Paid": (5, "Cash Flow"),
    "Cash Tax Paid": (5, "Cash Flow"),
    "Cash Pension Paid": (5, "Cash Flow"),
    "Cash Directors Bonus": (5, "Cash Flow"),
    "Capex Fixed Assets": (5, "Cash Flow"),
    "Capex Leased Assets": (5, "Cash Flow"),
    "Fixed Asset Sale Proceeds": (5, "Cash Flow"),
    "Dividends Received (CF)": (5, "Cash Flow"),
    "Securities Purchased": (5, "Cash Flow"),
    "Securities Sold": (5, "Cash Flow"),
    "Cash Dividends Paid": (5, "Cash Flow"),
    "Share Buyback": (5, "Cash Flow"),
    "Treasury Stock Sold": (5, "Cash Flow"),
    "Subordinated Bond Proceeds": (5, "Cash Flow"),
    "Paid-in Capital Received": (5, "Cash Flow"),
    # ── 6. Key Ratios (annual only) ─────────────────────────
    "Operating Cash Flow": (6, "Ratios"),
    "Free Cash Flow": (6, "Ratios"),
    "FCF per Share": (6, "Ratios"),
    "Capital expend": (6, "Ratios"),
    "BVPS": (6, "Ratios"),
    "Tangible BVPS": (6, "Ratios"),
    "PBV": (6, "Ratios"),
    "Price to Tangible BV": (6, "Ratios"),
    "EPS": (6, "Ratios"),
    "PE Ratio": (6, "Ratios"),
    "P/S": (6, "Ratios"),
    "EBITDA": (6, "Ratios"),
    "ROA": (6, "Ratios"),
    "ROE": (6, "Ratios"),
    "ROCE": (6, "Ratios"),
    "ROIC": (6, "Ratios"),
    "Interest Coverage": (6, "Ratios"),
    "Total Liab to Equity": (6, "Ratios"),
    "Debt to Equity": (6, "Ratios"),
    "LT Debt to Equity": (6, "Ratios"),
    "Debt to Assets": (6, "Ratios"),
    "Net Debt to Equity": (6, "Ratios"),
    "ST Debt": (6, "Ratios"),
    "LT Debt": (6, "Ratios"),
    "Total Debt": (6, "Ratios"),
    "Net Debt": (6, "Ratios"),
    "Working Capital": (6, "Ratios"),
    "Current Ratio": (6, "Ratios"),
    "Quick Ratio": (6, "Ratios"),
    "Financial Leverage": (6, "Ratios"),
    "Asset Turnover": (6, "Ratios"),
    "Fixed Asset Turnover": (6, "Ratios"),
    "Inventory Turnover": (6, "Ratios"),
    "Receivables Turnover": (6, "Ratios"),
    "DSO": (6, "Ratios"),
    "Days Inventory": (6, "Ratios"),
    "DPO": (6, "Ratios"),
    "Cash Conversion Cycle": (6, "Ratios"),
    "Working Capital Turnover": (6, "Ratios"),
    "Working Capital Ratio": (6, "Ratios"),
    # ── 7. Other (Dividend, etc.) ───────────────────────────
    "Dividend": (7, "Other"),
}


def classify(metric: str) -> tuple[int, str]:
    """Return (section_order, section_label) untuk metric."""
    # Strip "(Quarterly)" suffix for classification
    base = re.sub(r"\s*\(Quarterly\)\s*$", "", metric)
    return METRIC_SECTIONS.get(base, (99, "Other"))


def is_annual_period(p: str) -> bool:
    """True kalau period adalah tahun annual ('2025'), False kalau kuartal ('2025-Q1')."""
    return bool(re.match(r"^\d{4}$", p))


def period_sort_key(p: str) -> tuple[int, int]:
    """Sort 2008 < 2008-Q1 < 2008-Q2 < ... < 2025 < 2025-Q1."""
    try:
        if "-Q" in p:
            y, q = p.split("-Q", 1)
            return (int(y), int(q))
        return (int(p), 0)
    except (ValueError, IndexError):
        return (0, 0)


def collect(records):
    """Walk records, return:
      rows_annual: list of {ticker, name, sector, section, section_order, metric, values}
      rows_quarterly: same shape
      annual_periods, quarterly_periods: sorted lists
    """
    rows_annual = []
    rows_quarterly = []
    annual_periods = set()
    quarterly_periods = set()

    for rec in records:
        if not isinstance(rec, dict):
            continue
        ticker = rec.get("tick") or ""
        name = rec.get("name") or ""
        sector = rec.get("sector") or ""

        for k, v in rec.items():
            if k in META_FIELDS or k.startswith("_") or k in NON_TIMESERIES_FIELDS:
                continue
            if not isinstance(v, dict):
                continue

            # Split values into annual vs quarterly buckets
            annual_vals: dict[str, float] = {}
            q_vals: dict[str, float] = {}
            for p, val in v.items():
                if not isinstance(val, (int, float)) or val == 0:
                    continue
                if is_annual_period(p):
                    annual_vals[p] = val
                else:
                    q_vals[p] = val

            section_order, section_label = classify(k)

            if annual_vals:
                annual_periods.update(annual_vals.keys())
                rows_annual.append({
                    "ticker": ticker,
                    "name": name,
                    "sector": sector,
                    "section_order": section_order,
                    "section": section_label,
                    "metric": k,
                    "values": annual_vals,
                })
            if q_vals:
                quarterly_periods.update(q_vals.keys())
                rows_quarterly.append({
                    "ticker": ticker,
                    "name": name,
                    "sector": sector,
                    "section_order": section_order,
                    "section": section_label,
                    "metric": k.replace(" (Quarterly)", ""),
                    "values": q_vals,
                })

    # Sort rows by (ticker, section_order, metric)
    def row_sort_key(r):
        return (r["ticker"], r["section_order"], r["metric"])

    rows_annual.sort(key=row_sort_key)
    rows_quarterly.sort(key=row_sort_key)

    return (
        rows_annual,
        rows_quarterly,
        sorted(annual_periods, key=period_sort_key),
        sorted(quarterly_periods, key=period_sort_key),
    )


def write_csv(records, output_path):
    """CSV long format flat — tiap row 1 cell (ticker, metric, period, value)."""
    rows_annual, rows_quarterly, _, _ = collect(records)
    all_rows = rows_annual + rows_quarterly

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ticker", "name", "sector", "section", "metric", "period", "value"])
        for r in all_rows:
            for period, val in sorted(r["values"].items(), key=lambda x: period_sort_key(x[0])):
                w.writerow([
                    r["ticker"], r["name"], r["sector"],
                    r["section"], r["metric"], period, val,
                ])
    print(f"[OK] CSV written: {output_path} ({sum(len(r['values']) for r in all_rows)} value rows)")


def write_excel(records, output_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows_annual, rows_quarterly, annual_periods, quarterly_periods = collect(records)

    wb = openpyxl.Workbook()
    # Default sheet → Annual
    ws_annual = wb.active
    ws_annual.title = "Annual"
    _populate_sheet(ws_annual, rows_annual, annual_periods, get_column_letter)

    # Quarterly sheet
    ws_q = wb.create_sheet("Quarterly")
    _populate_sheet(ws_q, rows_quarterly, quarterly_periods, get_column_letter)

    wb.save(output_path)
    print(
        f"[OK] Excel written: {output_path} "
        f"(Annual: {len(rows_annual)} rows × {len(annual_periods)} periods, "
        f"Quarterly: {len(rows_quarterly)} rows × {len(quarterly_periods)} periods)"
    )


def _populate_sheet(ws, rows, periods, col_letter):
    from openpyxl.styles import Font, PatternFill, Alignment

    header = ["ticker", "name", "sector", "section", "metric", *periods]
    ws.append(header)

    bold_white = Font(bold=True, color="FFFFFF")
    dark_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    center = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = bold_white
        cell.fill = dark_fill
        cell.alignment = center

    section_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    section_font = Font(bold=True, color="60A5FA")  # blue-400

    current_section = None
    for r in rows:
        # Insert section header row when section changes (only if data has >1 ticker)
        if r["section"] != current_section:
            current_section = r["section"]
            # (skip section divider for now — keeps single-ticker exports clean)

        line = [r["ticker"], r["name"], r["sector"], r["section"], r["metric"]]
        for p in periods:
            line.append(r["values"].get(p, ""))
        ws.append(line)

    # Column widths
    widths = {"A": 10, "B": 36, "C": 22, "D": 18, "E": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # All period cols
    for i, _ in enumerate(periods, start=6):
        ws.column_dimensions[col_letter(i)].width = 14

    ws.freeze_panes = "F2"  # freeze 5 left cols + header row


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--format", required=True, choices=["excel", "csv"])
    args = p.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"[ERR] input not found: {src}")

    with open(src, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"[CONVERT] {len(data)} records from {src.name}")

    if args.format == "excel":
        write_excel(data, args.output)
    else:
        write_csv(data, args.output)


if __name__ == "__main__":
    main()
